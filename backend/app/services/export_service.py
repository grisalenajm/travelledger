import csv
import io
import logging
import re
import unicodedata
import zipfile
from decimal import Decimal
from pathlib import Path

import aiofiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.expense import Expense
from app.models.trip import Trip
from app.models.user import User
from app.services import paperless_service

logger = logging.getLogger(__name__)


def _slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-+", "-", text)
    return text.strip("-")


def _best_effort_filename(expense: Expense) -> str:
    """Returns a best-effort image filename for standalone CSV (no actual download)."""
    base = f"{expense.category.lower()}_{expense.date}_{expense.id}"
    if expense.local_path:
        ext = Path(expense.local_path).suffix or ".jpg"
        return f"{base}{ext}"
    if expense.paperless_doc_id:
        return f"{base}.pdf"
    return ""


async def build_csv(
    db: AsyncSession,
    trip: Trip,
    user: User,
    expenses: list[Expense],
    loyalty_cards: dict[str, str],
    image_map: dict[str, str] | None = None,
) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, delimiter=",", quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        "date", "description", "category", "billable", "payment_method",
        "loyalty_card", "amount", "currency", "amount_base", "base_currency",
        "exchange_rate", "rate_date", "paperless_url", "image_file",
    ])

    for expense in expenses:
        if image_map is not None:
            image_file = image_map.get(str(expense.id), "")
        else:
            image_file = _best_effort_filename(expense)

        loyalty = ""
        if expense.loyalty_card_id:
            loyalty = loyalty_cards.get(str(expense.loyalty_card_id), "")

        exchange_rate = ""
        if expense.amount and float(expense.amount) != 0:
            rate = float(expense.amount_base) / float(expense.amount)
            exchange_rate = f"{rate:.6f}"

        writer.writerow([
            str(expense.date),
            expense.description or "",
            expense.category,
            str(expense.billable).lower(),
            "",
            loyalty,
            str(expense.amount),
            expense.currency,
            str(expense.amount_base),
            user.currency_base,
            exchange_rate,
            str(expense.rate_date),
            "",
            image_file,
        ])

    return ("﻿" + output.getvalue()).encode("utf-8")


_EXT_MAP = {
    "image/jpeg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "application/pdf": ".pdf",
}


async def _collect_images(
    expenses: list[Expense],
    paperless_url: str | None,
    paperless_token: str | None,
) -> dict[str, tuple[str, bytes]]:
    """Returns {expense_id: (filename, bytes)} for expenses that have an image."""
    result: dict[str, tuple[str, bytes]] = {}

    for expense in expenses:
        str_id = str(expense.id)
        base = f"{expense.category.lower()}_{expense.date}_{str_id}"

        if expense.local_path:
            try:
                async with aiofiles.open(expense.local_path, "rb") as f:
                    content = await f.read()
                ext = Path(expense.local_path).suffix or ".jpg"
                result[str_id] = (f"{base}{ext}", content)
                logger.info("Export bundle — local image expense=%s", str_id)
            except Exception as e:
                logger.warning(
                    "Export bundle — fallo leyendo local_path=%s: %s",
                    expense.local_path, e,
                )

        elif expense.paperless_doc_id and paperless_url and paperless_token:
            try:
                image_bytes, content_type = await paperless_service.download_document(
                    paperless_url=paperless_url,
                    token=paperless_token,
                    doc_id=expense.paperless_doc_id,
                )
                ext = _EXT_MAP.get(content_type, ".bin")
                result[str_id] = (f"{base}{ext}", image_bytes)
                logger.info(
                    "Export bundle — Paperless doc_id=%s expense=%s",
                    expense.paperless_doc_id, str_id,
                )
            except Exception as e:
                logger.warning(
                    "Export bundle — fallo al descargar doc_id=%s: %s",
                    expense.paperless_doc_id, e,
                )

    return result


async def build_bundle(
    db: AsyncSession,
    trip: Trip,
    user: User,
    expenses: list[Expense],
    loyalty_cards: dict[str, str],
    format: str = "xlsx",
    payment_methods: dict[str, str] | None = None,
) -> bytes:
    paperless_url, paperless_token = await paperless_service.get_credentials(db, user.id)
    images = await _collect_images(expenses, paperless_url, paperless_token)

    image_map = {eid: fname for eid, (fname, _) in images.items()}
    trip_slug = _slugify(trip.name)

    if format == "xlsx":
        data_bytes = generate_xlsx(expenses, trip, user, payment_methods)
        data_filename = f"gastos_{trip_slug}.xlsx"
    else:
        data_bytes = await build_csv(db, trip, user, expenses, loyalty_cards, image_map)
        data_filename = f"gastos_{trip_slug}.csv"

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(data_filename, data_bytes)
        for filename, content in images.values():
            zf.writestr(filename, content)

    zip_buffer.seek(0)
    return zip_buffer.read()


def _fmt_decimal_eu(value: Decimal | float) -> str:
    """Formats number with comma as decimal separator (European style): 1.234,56"""
    formatted = f"{float(value):,.2f}"
    return formatted.replace(",", "X").replace(".", ",").replace("X", ".")


def _payment_method_name(exp: Expense, payment_methods: dict[str, str] | None) -> str:
    if exp.payment_method_id and payment_methods:
        return payment_methods.get(str(exp.payment_method_id), "")
    return ""


def generate_csv(
    expenses: list[Expense],
    trip: Trip,
    user: User,
    payment_methods: dict[str, str] | None = None,
) -> bytes:
    """
    European CSV: field separator=';', decimal=','.
    UTF-8 BOM so Excel in Spanish opens without conversion.
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_MINIMAL)

    writer.writerow([
        "Fecha", "Descripción", "Categoría",
        "Moneda gasto", "Importe gasto",
        "Moneda base", f"Importe base ({user.currency_base})",
        "Facturable", "Método de pago",
    ])

    currency_totals: dict[str, Decimal] = {}
    total_base = Decimal("0")

    for exp in expenses:
        writer.writerow([
            exp.date.isoformat(),
            exp.description or "",
            exp.category,
            exp.currency,
            _fmt_decimal_eu(exp.amount),
            user.currency_base,
            _fmt_decimal_eu(exp.amount_base),
            "Sí" if exp.billable else "No",
            _payment_method_name(exp, payment_methods),
        ])
        currency_totals[exp.currency] = (
            currency_totals.get(exp.currency, Decimal("0")) + Decimal(str(exp.amount))
        )
        total_base += Decimal(str(exp.amount_base))

    writer.writerow([])
    writer.writerow(["TOTALES POR MONEDA DEL GASTO"])
    for currency, total in sorted(currency_totals.items()):
        writer.writerow(["", currency, _fmt_decimal_eu(total)])

    writer.writerow([])
    writer.writerow([f"TOTAL EN {user.currency_base}", "", _fmt_decimal_eu(total_base)])

    return ("﻿" + output.getvalue()).encode("utf-8")


def generate_xlsx(
    expenses: list[Expense],
    trip: Trip,
    user: User,
    payment_methods: dict[str, str] | None = None,
) -> bytes:
    wb = Workbook()

    # ── Sheet 1: Gastos ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Gastos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="004D64")
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    HEADERS = [
        "Fecha", "Descripción", "Categoría",
        "Moneda gasto", "Importe gasto",
        "Moneda base", "Importe base",
        "Facturable", "Método de pago",
    ]
    COL_WIDTHS = [12, 35, 15, 14, 16, 12, 16, 12, 18]

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    currency_totals: dict[str, Decimal] = {}
    total_base = Decimal("0")

    for row_num, exp in enumerate(expenses, start=2):
        ws.cell(row=row_num, column=1, value=exp.date.isoformat())
        ws.cell(row=row_num, column=2, value=exp.description or "")
        ws.cell(row=row_num, column=3, value=exp.category)
        ws.cell(row=row_num, column=4, value=exp.currency)
        amount_cell = ws.cell(row=row_num, column=5, value=float(exp.amount))
        amount_cell.number_format = "#,##0.00"
        ws.cell(row=row_num, column=6, value=user.currency_base)
        base_cell = ws.cell(row=row_num, column=7, value=float(exp.amount_base))
        base_cell.number_format = "#,##0.00"
        ws.cell(row=row_num, column=8, value="Sí" if exp.billable else "No")
        ws.cell(row=row_num, column=9, value=_payment_method_name(exp, payment_methods))

        currency_totals[exp.currency] = (
            currency_totals.get(exp.currency, Decimal("0")) + Decimal(str(exp.amount))
        )
        total_base += Decimal(str(exp.amount_base))

    summary_start = len(expenses) + 3

    ws.cell(row=summary_start, column=1, value="TOTALES POR MONEDA DEL GASTO").font = bold
    for i, (currency, total) in enumerate(sorted(currency_totals.items())):
        row = summary_start + 1 + i
        ws.cell(row=row, column=4, value=currency)
        tc = ws.cell(row=row, column=5, value=float(total))
        tc.number_format = "#,##0.00"
        tc.font = bold

    total_row = summary_start + len(currency_totals) + 2
    label = ws.cell(row=total_row, column=1, value=f"TOTAL EN {user.currency_base}")
    label.font = Font(bold=True, color="FFFFFF")
    label.fill = header_fill
    tc = ws.cell(row=total_row, column=7, value=float(total_base))
    tc.number_format = "#,##0.00"
    tc.font = Font(bold=True, color="FFFFFF")
    tc.fill = header_fill

    # ── Sheet 2: Resumen ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen")
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 18

    ws2.cell(row=1, column=1, value="Viaje").font = bold
    ws2.cell(row=1, column=2, value=trip.name)
    ws2.cell(row=2, column=1, value="Período").font = bold
    ws2.cell(row=2, column=2, value=f"{trip.start_date} → {trip.end_date}")
    ws2.cell(row=3, column=1, value="Moneda base").font = bold
    ws2.cell(row=3, column=2, value=user.currency_base)

    counts: dict[str, int] = {}
    for exp in expenses:
        counts[exp.currency] = counts.get(exp.currency, 0) + 1

    ws2.cell(row=5, column=1, value="Por moneda").font = Font(bold=True)
    ws2.cell(row=5, column=2, value="Gastos").font = Font(bold=True)
    ws2.cell(row=5, column=3, value="Total").font = Font(bold=True)

    for i, (currency, total) in enumerate(sorted(currency_totals.items())):
        row = 6 + i
        ws2.cell(row=row, column=1, value=currency)
        ws2.cell(row=row, column=2, value=counts.get(currency, 0))
        tc = ws2.cell(row=row, column=3, value=float(total))
        tc.number_format = "#,##0.00"

    base_row = 6 + len(currency_totals) + 1
    ws2.cell(row=base_row, column=1, value=f"Total {user.currency_base}").font = bold
    ws2.cell(row=base_row, column=2, value=len(expenses))
    tbc = ws2.cell(row=base_row, column=3, value=float(total_base))
    tbc.number_format = "#,##0.00"
    tbc.font = bold

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
